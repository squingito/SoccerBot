from openpyxl import load_workbook

workbook = load_workbook(filename='PL_dataset.xlsx')
workbook.sheetnames

sheet = workbook.active
getStatsArray = ["F", "E", "G", "J", "AC", "AL", "AM", "AD", "AN"]
defender =  ["N", "O", "P", "Q", "R", "S", "T", "AA", "AB"]
midfielder = ["P", "Q", "AH", "S" "U", "AE", "AF", "AI", "AO"]
goalkeeper = ["N", "O", "AJ", "AK"]
forward = ["K", "L", "M", "S", "AE", "AF", "AO"]


#sheet["A"]
def findIndexOf(name):
    i = 1
    valueFound = False
    for value in sheet.iter_rows(min_row = 1, max_row = 572,
                                 min_col = 1, max_col = 1,
                                 values_only = True):
        tempValue = value[0]
        if (name.lower() == tempValue.lower()):
            rowVar1 = i
            valueFound = True
            break
        i = i + 1
    if valueFound:
        return i
    else:
        return None

def getStats(name):
    if(len(name.split(" ")) >= 2):
        name = name[10:]
        print(name)
        if(name == "clown"):
            print(name)
            name = "Granit Xhaka"
        if(name == "goat"):
            name = "Harry Kane"

        statArray = None
        a1 = ""
        a2 = ""
        num = findIndexOf(name)
        formatType = sheet["D" + str(num)].value
        if(num != None):
            for each in getStatsArray:
                a1 = a1 + (str(sheet[each+str(1)].value) + ":\n")
                a2 = a2 + (str(sheet[each+str(num)].value) +"\n")
            if formatType == "Goalkeeper":
                statArray = goalkeeper
                pos = "Goalkeeper"
            elif formatType == "Defender":
                statArray = defender
                pos = "Defender"
            elif formatType == "Forward":
                statArray = forward
                pos = "Forward"
            elif formatType == "Midfielder":
                statArray = midfielder
                pos = "Midfieler"

            if(statArray != None):
                for each in statArray:
                    a1 = a1 + (str(sheet[each+str(1)].value) + ":\n")
                    a2 = a2 + (str(sheet[each+str(num)].value) +"\n")
                return a1,a2, (str(sheet["A" + str(num)].value))
        else:
            return '"' + name + '"' + " does not exist",None,None
    else:
        return "please write a person after the space",None,None

def compare(names):
    name1 = names.split("|")[0][9:]
    name2 = names.split("|")[1]
    while(1):
        if name1[-1] == " ":
            name1 = name1[:-1]
        else:
            break
    while(1):
        if name2[0] == " ":
            name2 = name2[1:]
        else:
            break
    text = ""
    formatType = None
    num1 = findIndexOf(name1)
    num2 = findIndexOf(name2)
    statArray = None
    a1 = ""
    a2 = ""
    a3 = ""
    pos = None
    
    if num1 != None and num2 != None:
        if(sheet["D" + str(num1)].value == sheet["D" + str(num2)].value):
            formatType = sheet["D" + str(num1)].value
        for each in getStatsArray:
            a1 = a1 + (str(sheet[each+str(1)].value) + ":\n")
            a2 = a2 + (str(sheet[each+str(num1)].value) +"\n")
            a3 = a3 + (str(sheet[each+str(num2)].value) + "\n")
        if formatType == "Goalkeeper":
            statArray = goalkeeper
            pos = "Goalkeeper"
        elif formatType == "Defender":
            statArray = defender
            pos = "Defender"
        elif formatType == "Forward":
            statArray = forward
            pos = "Forward"
        elif formatType == "Midfielder":
            statArray = midfielder
            pos = "Midfieler"

        if(statArray != None):
            for each in statArray:
                a1 = a1 + (str(sheet[each+str(1)].value) + ":\n")
                a2 = a2 + (str(sheet[each+str(num1)].value) +"\n")
                a3 = a3 + (str(sheet[each+str(num2)].value) + "\n")

        if(pos != None):
            return a1,a2,a3,pos,name1,name2
        else:
            return a1,a2,a3, None,name1,name2

        return formatGetStats(num1) + "\n" + "\n" + formatGetStats(num2)
    elif num1 == None and num2 == None:
        return '"' + name1 + '"'+ ' and ' +  '"' + name2 + '"'  + " do not exist",None,None,None,None,None
    elif num1 == None:
        return '"' + name1 + '"' + " does not exist",None,None,None,None,None
    else:
        return '"' + name2 + '"' + " does not exist",None,None,None,None,None
        


def formatGetStats(index):
    return "Name: {}\nGoals: {}\nAssists: {}\nYellow Cards: {}\nRed Cards: {}\nPasses: {}\nFouls: {}\nOffsides: {}\nAppearances: {}\nAge: {}"\
        .format(sheet["A"+ str(index)].value,sheet["J"+ str(index)].value, \
        sheet["AC"+ str(index)].value,sheet["AL"+ str(index)].value, sheet["AM"+ str(index)].value, sheet["AD"+ str(index)].value, \
        sheet["AN" + str(index)].value, sheet["AO" + str(index)].value, sheet["G" + str(index)].value, sheet["F" + str(index)].value) 

def forwardFormatGetStats(index):
    return "Shots on Target: {}\nShooting Accuracy (%): {}\nAerial Battles Won: {}\nAerial Battles Lost: {}\nBig Chances Created: {}\nCrosses: {}".format(sheet["L"+ str(index)].value,sheet["M"+ str(index)].value,sheet["Y"+ str(index)].value,sheet["Z"+ str(index)].value,sheet["AE"+ str(index)].value,sheet["AF"+ str(index)].value)
    
def midfieldFormatGetStats(index):
    return "Tackles: {}\nTackle Success (%): {}\nThrough Balls: {}\nInterceptions: {}\nRecoveries: {}\nBig Chances Created: {}\nCrosses: {}\nAccurate Long Balls: {}".format(sheet["P" + str(index)].value,sheet["Q" + str(index)].value,sheet["AH" + str(index)].value,sheet["S" + str(index)].value,sheet["U" + str(index)].value,sheet["AE" + str(index)].value,sheet["AF" + str(index)].value,sheet["AI" + str(index)].value)

def defenderFormatGetStats(index):
    return "Clean Sheets {}\nGoals Conceded {}\nTackles {}\nTackle Success % {}\nBlocked Shots {}\nInterceptions {}\nClearances {}\nOwn Goals {}\nErrors Leading to Goal {}"\
    .format(sheet["N" + str(index)], sheet["O" + str(index)], sheet["P" + str(index)], sheet["Q" + str(index)],\
    sheet["R" + str(index)], sheet["S" + str(index)], sheet["T" + str(index)], sheet["AA" + str(index)], sheet["AB" + str(index)])

def goalkeeperFormatGetStats(index):
    return "Clean Sheets {}\nGoals Conceded {}\nSaves {}\nPenalties Saved {}"\
    .format(sheet["N" + str(index)], sheet["O" + str(index)], sheet["AJ" + str(index)], sheet["AK" + str(index)])